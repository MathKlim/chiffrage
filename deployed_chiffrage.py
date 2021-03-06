import base64

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import save_virtual_workbook

frames = []
total_jours = []
total_paliers = []
total_equip = []
cout_equip = []
total_echeance = []

st.sidebar.image("logo/Logo_CITC_Gris.png", use_column_width=True)

paliers = st.sidebar.number_input("Nombre de paliers", value=1)
tjm = st.sidebar.number_input(
    "Cout mix ingénieur Sénior/Junior (TJH Moyen)", value=550
)

pilotage = st.sidebar.number_input("Coefficient pilotage", value=0.12)
risque = st.sidebar.number_input("Coefficient risque", value=0.20)

test_validation = st.sidebar.number_input(
    "Coefficient test et validation", value=0.20
)

integration = st.sidebar.number_input("Coefficient intégration", value=0.10)

add_pilotage = st.sidebar.checkbox(
    "Prendre en compte le pilotage dans le risque ? Non par défaut."
)
st.title("Simulateur de coût PTF")


st.markdown("# Paliers")


# definition du nombre de paliers
for palier in range(paliers):
    st.header(f"Palier {palier}")

    nom = st.text_input("Dénomination du palier", "palier", key=f"{palier}")

    # nombre de taches dans chaque palier
    taches = st.number_input(
        f"Nombre de tâches à effectuer dans le palier {palier}",
        value=0,
        key=f"{palier}",
    )

    subtotal_jours = []
    subtotal_tasks = []

    # definition du nom de la tache et du nombre de jours
    # estimé nécessaire à la tâche
    for tache in range(taches):
        left, right = st.beta_columns(2)
        with left:
            task = st.text_input(
                f"Tâche {tache}", "task", key=f"{palier}_{tache}"
            )
        with right:
            jours = st.number_input(
                "Nombre de jours estimés",
                max_value=100,
                min_value=0,
                value=0,
                key=f"{palier}_{tache}",
            )

        # chaque tache est ajouté à la liste liste pour
        # avoir un sous total par palier
        subtotal_tasks.append(
            {
                "palier": palier,
                "nom": nom,
                "tache": task,
                "jours": jours,
            }
        )

        # liste des nombres de jours par taches
        subtotal_jours.append(jours)

    # recap des taches pour chaque palier
    st.markdown("### Récapitulatif du palier")
    sub_descriptif = pd.DataFrame(
        subtotal_tasks, columns=["palier", "nom", "tache", "jours"]
    )
    st.table(sub_descriptif)
    frames.append(sub_descriptif)

    # calcul du nombre global de jour par palier
    total = np.sum(subtotal_jours)
    st.sidebar.write(
        f"Total de jours hommes pour le palier {palier} : ", total
    )

    total_jours.append(total)
    # total_paliers.append(subtotal_tasks)

# recap global des taches et paliers
st.markdown("### Descriptif global des paliers")
descriptif_global = pd.concat(frames, ignore_index=True)
st.table(descriptif_global)

total_jours = np.sum(total_jours)
st.sidebar.write("Nombre de jours au total pour le projet : ", total_jours)


nb_pilotage = float(total_jours * pilotage)
nb_test_validation = float(total_jours * test_validation)
nb_integration = float(total_jours * integration)

if add_pilotage:
    nb_risque = float((total_jours + nb_pilotage) * risque)
else:
    nb_risque = float(total_jours * risque)

total_jour_projet = float(
    total_jours + nb_pilotage + nb_risque + nb_test_validation + nb_integration
)

st.sidebar.write("Total de jours pour le pilotage : ", nb_pilotage)
st.sidebar.write("Total de jours pour le risque : ", nb_risque)
st.sidebar.write(
    "Total de jours pour les tests & validations : ", nb_test_validation
)
st.sidebar.write("Total de jours pour l'intégration : ", nb_integration)
st.sidebar.write(
    "Total de jours pour le projet, pilotage, risque, tests & integration compris : ",
    total_jour_projet,
)

cout_global_palier = float(total_jour_projet * tjm)

st.sidebar.write("Coût total jour homme : ", total_jours * tjm)

st.sidebar.write(
    "Coût total des paliers pour le projet, pilotage, risque, tests & intégration compris : ",
    cout_global_palier,
)

details = {
    "Nombre de jours homme au total pour le projet": total_jours,
    "Coût total jour homme": total_jours * tjm,
    "Coefficient pilotage": pilotage,
    "Coefficient risque": risque,
    "Coefficient test et validation": test_validation,
    "Coefficient intégration": integration,
    "Total de jours pour le pilotage": nb_pilotage,
    "Total de jours pour le risque ": nb_risque,
    "Total de jours pour les tests & validations": nb_test_validation,
    "Total de jours pour l'intégration": nb_integration,
    "Total de jours total pour le projet": total_jour_projet,
    "Coût total des paliers pour le projet, risque et pilotage compris ": cout_global_palier,
}
details = pd.DataFrame.from_dict(details, orient="index")

st.markdown("# Transverse")

transverse = []

st.markdown("## Documentation")

manuel_utilisateur = st.number_input(
    "Rédaction du manuel utilisateur", value=0
)

dossier_conception_deploiement = st.number_input(
    "Rédaction du dossier de conception et déploiement", value=0
)

plan_de_test = st.number_input("Rédaction plan de test d'intégration", value=0)
transverse.append(
    manuel_utilisateur + dossier_conception_deploiement + plan_de_test
)

st.markdown("## Maintient en condition opérationnel")

correctif = st.number_input(
    "Accompagnement sur les correctifs à mettre en place", value=0
)

st.markdown("## Reversibilité")

formation_reversibilite = st.number_input(
    "Formation suite à la reversibilité", value=0
)
transverse.append((formation_reversibilite + correctif) * pilotage)

acc_reversibilite = st.number_input(
    "Accompapgnement suite à la reversibilité", value=0
)
transverse.append(acc_reversibilite * pilotage)

st.markdown("## Workshop")

reunions = st.number_input(
    "Réunions de réflexion sur les différents sujets du projet", value=0
)
transverse.append(reunions)

st.markdown("## Adaptabilité expérience utilisateur")

ajustement = st.number_input(
    "Ajustements d'applications, ajout de fonctionnalités", value=0
)
transverse.append(ajustement)

st.markdown("## Formations")

utilisateur = st.number_input("Formations utilisateurs", value=0)
personnel = st.number_input("Formations personnels", value=0)
doc_formation = st.number_input("Documents formations", value=0)
transverse.append(utilisateur + personnel + doc_formation)

total_transverse = np.sum(transverse)
st.info(
    f"Nombre total de jours pour les activités transverse : {total_transverse}"
)

cout_global_transverse = total_transverse * tjm
st.sidebar.write(
    "Coût total des activités transverses : ",
    cout_global_transverse,
)
st.markdown("# Equipement")

nb_equip = st.number_input(
    "Nombre d'équipements à lister",
    value=1,
)

for equip in range(nb_equip):
    left, right = st.beta_columns(2)
    with left:
        equipement = st.text_input(
            f"Equipement {equip}", "equip", key=f"{equip}"
        )
    with right:
        cost = st.number_input(
            "Coût estimé",
            max_value=100000,
            min_value=0,
            value=0,
            key=f"{equip}",
        )
    cout_equip.append(cost)

    # chaque equipement est ajouté à la liste liste pour
    # avoir un sous total par palier
    total_equip.append({"equipement": equipement, "cout": cost})

st.markdown("### Récapitulatif équipement")
descriptif_equip = pd.DataFrame(total_equip, columns=["equipement", "cout"])
st.table(descriptif_equip)

# cout_equip
cout_global_equipement = np.sum(cout_equip)
st.sidebar.write(
    "Coût total estimé de l'équipement : ", cout_global_equipement
)

st.markdown("# Coût final")

cout_global = (
    cout_global_palier + cout_global_transverse + cout_global_equipement
)

st.markdown(
    f"## {float(cout_global):.2f} Euros",
)


st.markdown("# Echéancier de paiement")
st.markdown(
    "## **ATTENTION** : Par défaut, l'échéancier de paiement ne prend en compte"
    " que le coût total pour le projet. Le coût de l'équipement n'est pas pris en compte."
)
adhesion = st.number_input(
    "Coût de l'adhésion", min_value=0, max_value=10000, value=2200
)

add_adhesion = st.checkbox("Prendre en compte l'adhésion dans le coût global")

if add_adhesion:
    cout_global += adhesion

# nombre d'echeances voulues
echeances = st.number_input(
    "Nombre d'échéances",
    value=0,
)

# definition du nom de la tache et du nombre de jours
# estimé nécessaire à la tâche
for echeance in range(echeances):
    left, center, right = st.beta_columns(3)
    with left:
        ratio = st.number_input(
            f"Ratio a payer à l'échéance {echeance}",
            min_value=0.0,
            max_value=1.0,
            value=0.3,
            key=f"{echeance}",
        )
    with center:
        descriptif = st.text_input(
            "Descriptif", "description", key=f"{echeance}"
        )
    with right:
        montant = st.text_input(
            "Montant en Euro HT", ratio * cout_global, key=f"{echeance}"
        )

    # chaque echeance est ajouté à la liste liste pour
    # avoir un total
    total_echeance.append(
        {"ratio": ratio, "descriptif": descriptif, "montant HT": montant}
    )

# recap des taches pour chaque palier
st.markdown("### Récapitulatif des écheances de paiement")
echeancier = pd.DataFrame(
    total_echeance, columns=["ratio", "descriptif", "montant HT"]
)
st.table(echeancier)


def to_excel(dfs):
    wb = openpyxl.Workbook()
    # ws = wb.active
    titre_feuille = [
        "Paliers",
        "Details",
        "Equipement",
        "Echeancier",
    ]
    for i, df in enumerate(dfs):
        wb.create_sheet(index=int(i), title=f"{titre_feuille[i]}")
        ws = wb[f"{titre_feuille[i]}"]
        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        for cell in ws["A"] + ws[1]:
            cell.style = "Pandas"
    stream = save_virtual_workbook(wb)
    return stream


def get_table_download_link(df):
    """Generates a link allowing the data in a given panda dataframe to be downloaded
    in:  dataframe
    out: href string
    """
    val = to_excel(df)
    b64 = base64.b64encode(
        val
    ).decode()  # some strings <-> bytes conversions necessary here
    href = f'<a href="data:file/csv;base64,{b64}" download="PTF.xlsx">Télécharger la PTF</a>'
    return href


dfs = [descriptif_global, details, descriptif_equip, echeancier]

dl = st.button("Téléchargement")

if dl:
    st.markdown(get_table_download_link(dfs), unsafe_allow_html=True)
